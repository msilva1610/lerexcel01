import openpyxl

import glob, os
import os.path
import myutils
import datetime
from openpyxl import load_workbook
import config as cfg

def read_excel(path):
    """
    Recebe o caminho completo do excel, carrega o excel e vai para a pasta a ser lida
    """
    wb = openpyxl.load_workbook(path, read_only = True)
    ws_Projetos = wb[cfg.workbook]
    total_linhas = ws_Projetos.max_row
    print("Total de Linhas na aba projetos: {}".format(total_linhas))

    listaDeProjetos = []
    for linha in range(cfg.linhaInicial, total_linhas+1):
        projetos = {}
        if ws_Projetos.cell(row=linha, column=cfg.excelColumn['RM']).value == None:
            print("Ultima linha lida: {}".format(linha-1))
            break
        else:
            projetos.update({'Tipo': str (ws_Projetos.cell(row=linha, column=cfg.excelColumn['TIPO']).value) })
            projetos.update({'Versão': str (ws_Projetos.cell(row=linha, column=cfg.excelColumn['Versão']).value) })
            projetos.update({'RM': str (ws_Projetos.cell(row=linha, column=cfg.excelColumn['RM']).value) })
            projetos.update({'Revisao':  ws_Projetos.cell(row=linha, column=cfg.excelColumn['Revisao']).value})
            projetos.update({'Data de Liberação': str (ws_Projetos.cell(row=linha, column=cfg.excelColumn['Data de Liberação']).value) })
            projetos.update({'Data de Inicio':  str (ws_Projetos.cell(row=linha, column=cfg.excelColumn['Data de Inicio']).value)})
            projetos.update({'Término da Execução':  str (ws_Projetos.cell(row=linha, column=cfg.excelColumn['Término da Execução']).value)})
            projetos.update({'Resultado':  str (ws_Projetos.cell(row=linha, column=cfg.excelColumn['Resultado']).value)})
            projetos.update({'Descrição':  str (ws_Projetos.cell(row=linha, column=cfg.excelColumn['Descrição']).value)})
            projetos.update({'Tecnologia':  str (ws_Projetos.cell(row=linha, column=cfg.excelColumn['Tecnologia']).value)})
            projetos.update({'Passos':  str (ws_Projetos.cell(row=linha, column=cfg.excelColumn['Passos']).value)})
            projetos.update({'Ambiente Suportado':  str (ws_Projetos.cell(row=linha, column=cfg.excelColumn['Ambiente Suportado']).value)})
            projetos.update({'Origem RM':  str (ws_Projetos.cell(row=linha, column=cfg.excelColumn['Origem RM']).value)})
            projetos.update({'NOME PROJETO / NUMERO INCIDENTE':  str (ws_Projetos.cell(row=linha, column=cfg.excelColumn['NOME PROJETO / NUMERO INCIDENTE']).value)})
            projetos.update({'Tempo execução RM':  str (ws_Projetos.cell(row=linha, column=cfg.excelColumn['Tempo execução RM']).value)})

        listaDeProjetos.append(projetos)
        # if linha == 500:
        #     break
    return listaDeProjetos

def main():
    os.chdir(cfg.origemArquivosExcel) # Alterar diretório de trabalho
    for i in os.listdir(os.getcwd()):
        # print(i)
        print ("Lendo arquivo: {}".format(os.path.join(os.getcwd(), i)))
        if os.path.isfile(i):
            # print(os.path.isfile(i))
            if i.endswith(".xlsx"):
                DataCriacaoDoArquivo = myutils.creation_date(os.path.join(os.getcwd(), i))
                print("Data criação do arquivo: {}".format(datetime.datetime.fromtimestamp(DataCriacaoDoArquivo)))
                pathcompleto = os.path.join(os.getcwd(), i)
                listaDeRmsAplicadas = read_excel(pathcompleto)
                myutils.salvaProjetos(listaDeRmsAplicadas)
                print("Arquivo json gerado")

if __name__ == "__main__":
	main()




