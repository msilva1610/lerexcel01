import glob, os
import os.path
import datetime
import shutil
import platform
import io
import json
import datetime 
# from dateutil.parser import parse

from openpyxl import load_workbook
import config as cfg

def salvaProjetos(lista):
    with io.open('ControleDeRMsAmbientes.json', 'w', encoding='utf8', errors='ignore') as outfile:  
        json.dump(lista, outfile, ensure_ascii=False)


def readFilesFromSource():
    """
    Ler o arquivo no formato Excel
    """
    os.chdir(cfg.origemArquivosExcel)

    projetos = {}
    listaDeProjetos = []

    for i in os.listdir(os.getcwd()):
        print (os.path.join(os.getcwd(), i))
        print(i)
        if os.path.isfile(i):
            # if i.endswith(".xls") or i.endswith(".xlsx") or i.endswith(".xlsm"):
            if i.endswith(".xlsx"):
                DataCriacaoDoArquivo = creation_date(os.path.join(os.getcwd(), i))
                print ("Data criação do Arquivo: {}".format(datetime.datetime.fromtimestamp(DataCriacaoDoArquivo)))
                NomeDoArquivo = (os.path.join(os.getcwd(), i))
                NomeDoArquivo1 = (os.path.basename(NomeDoArquivo))
                print("Filename: {}".format(NomeDoArquivo))
                print("Filename with path: {}".format(NomeDoArquivo1))
                caminho = os.path.join(os.getcwd())
                wb = load_workbook(caminho, i)

                                
def carregarExcel(caminho,i):
    """
    Ler arquivo no formato Excel e devolve um json file
    """
    wb = load_workbook(caminho, i)
    # ws_Projetos = wb[cfg.workbook]
    # total_linhas = ws_Projetos.max_row
    # print("Total de linhas: {}".format(total_linhas))
    # return {'Total de linhas': total_linhas}


def main():
    readFilesFromSource()

if __name__ == "__main__":
	main()